"""
Extract data from Xilnex Portal Excel export and convert to JSON format
Matches the structure of my_portal.json for comparison

Usage:
    python extract_xilnex_excel.py
"""
import json
import re
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def parse_currency(value):
    """
    Parse currency string to float
    Handles formats like:
    - "RM 1,234.56"
    - "RM1,234.56"
    - "1,234.56"
    - "(RM 100.00)" for negatives
    - "-RM 100.00" for negatives
    """
    if value is None:
        return 0.0
    
    # If already a number, return it
    if isinstance(value, (int, float)):
        return float(value)
    
    # Convert to string
    value_str = str(value).strip()
    
    # Check for negative (parentheses or minus sign)
    is_negative = False
    if value_str.startswith('(') and value_str.endswith(')'):
        is_negative = True
        value_str = value_str[1:-1]  # Remove parentheses
    elif value_str.startswith('-'):
        is_negative = True
        value_str = value_str[1:]  # Remove minus
    
    # Remove "RM" prefix if present
    value_str = re.sub(r'^RM\s*', '', value_str, flags=re.IGNORECASE)
    
    # Remove commas and convert to float
    value_str = value_str.replace(',', '')
    
    try:
        result = float(value_str)
        return -result if is_negative else result
    except (ValueError, TypeError):
        return 0.0

def normalize_store_name(store_name):
    """Normalize store name: trim whitespace, uppercase"""
    if store_name is None:
        return None
    return str(store_name).strip().upper()

def parse_date(date_value):
    """
    Parse date from various formats to YYYY-MM-DD
    Handles Excel date numbers, datetime objects, and string dates
    """
    if date_value is None:
        return None
    
    # If it's already a datetime object
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    
    # If it's an Excel date number (openpyxl handles this)
    if isinstance(date_value, (int, float)):
        # Try to convert Excel serial date
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(date_value)
            return dt.strftime('%Y-%m-%d')
        except:
            pass
    
    # If it's a string, try to parse
    if isinstance(date_value, str):
        # Try common date formats
        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(date_value.strip(), fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
    
    return None

def find_header_row(ws):
    """
    Find the row containing headers
    Looks for "Store", "Date", "Grand Total" or similar headers
    Skips title rows (rows with "DAILY SALES" or similar)
    """
    header_keywords = ['store', 'date', 'grand total', 'sales', 'profit']
    skip_keywords = ['daily sales', 'summary', 'generated', 'filter']
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        row_str = ' '.join(str(cell).lower() if cell else '' for cell in row)
        
        # Skip title/metadata rows
        if any(skip_kw in row_str for skip_kw in skip_keywords):
            continue
        
        # Check if this looks like a header row
        if any(keyword in row_str for keyword in header_keywords):
            # Make sure it's not just a title row
            if 'store' in row_str.lower() and 'date' in row_str.lower():
                return row_idx
    
    # Default to row 10 if not found (based on portal export structure)
    return 10

def extract_xilnex_data(excel_path):
    """
    Extract data from Xilnex Excel file
    Returns list of dicts with: date, store_name, sales_amount, profit_amount
    """
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # Find header row
    header_row = find_header_row(ws)
    print(f"Found headers at row {header_row}")
    
    # Read header row to identify column positions
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value).strip().upper() if cell.value else '')
    
    print(f"Headers at row {header_row}: {headers}")
    
    # Also check a few rows below to see the data structure
    print("\nSample data rows (next 3 rows after header):")
    for i in range(1, 4):
        row_num = header_row + i
        sample_row = []
        for cell in ws[row_num]:
            sample_row.append(str(cell.value) if cell.value is not None else '')
        print(f"  Row {row_num}: {sample_row[:5]}")  # First 5 columns
    
    # Identify column indices
    # Expected columns: Store, Date, Sales (Grand Total), Profit
    store_col = None
    date_col = None
    sales_col = None
    profit_col = None
    
    for idx, header in enumerate(headers):
        header_upper = header.upper()
        if 'STORE' in header_upper and store_col is None:
            store_col = idx
        elif 'DATE' in header_upper and date_col is None:
            date_col = idx
        elif ('SALES' in header_upper or 'GRAND TOTAL' in header_upper or 'TOTAL' in header_upper) and sales_col is None:
            # Check if it's not profit
            if 'PROFIT' not in header_upper:
                sales_col = idx
        elif 'PROFIT' in header_upper and profit_col is None:
            profit_col = idx
    
    # If columns not found by header, try positional (A=Store, B=Date, C=Sales, D=Profit)
    # But first check if we can infer from data
    if store_col is None or date_col is None:
        # Look at first data row to infer structure
        first_data_row = header_row + 1
        if first_data_row <= ws.max_row:
            first_row_values = [cell.value for cell in ws[first_data_row]]
            # Column A usually has store names (text), Column B has dates
            if len(first_row_values) > 0:
                if store_col is None and isinstance(first_row_values[0], str) and 'MB' in first_row_values[0].upper():
                    store_col = 0
                if date_col is None and len(first_row_values) > 1:
                    date_col = 1
    
    # Default positions if still not found
    if store_col is None:
        store_col = 0  # Column A
    if date_col is None:
        date_col = 1  # Column B
    if sales_col is None:
        sales_col = 2  # Column C
    if profit_col is None:
        profit_col = 3  # Column D
    
    print(f"Column mapping: Store={store_col}, Date={date_col}, Sales={sales_col}, Profit={profit_col}")
    
    # Extract data rows (start after header row)
    data = []
    data_start_row = header_row + 1
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # Skip empty rows
        if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
            continue
        
        # Skip total rows (rows that start with "Total" or are all empty except totals)
        if row[store_col] and isinstance(row[store_col], str):
            if row[store_col].strip().upper().startswith('TOTAL'):
                continue
        
        store_name = normalize_store_name(row[store_col] if store_col < len(row) else None)
        date_str = parse_date(row[date_col] if date_col < len(row) else None)
        sales_amount = parse_currency(row[sales_col] if sales_col < len(row) else None)
        profit_amount = parse_currency(row[profit_col] if profit_col < len(row) else None)
        
        # Skip rows without store name or date
        if not store_name or not date_str:
            continue
        
        data.append({
            'date': date_str,
            'store_name': store_name,
            'sales_amount': round(sales_amount, 2),
            'profit_amount': round(profit_amount, 2)
        })
    
    print(f"Extracted {len(data)} records")
    return data

def main():
    excel_path = Path(r"C:\Users\yongw\Downloads\Daily Sales (Summary)_26-11-2025 (Web).xlsx")
    output_path = Path(__file__).parent / "xilnex_portal.json"
    
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return
    
    # Extract data
    data = extract_xilnex_data(excel_path)
    
    # Save to JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"\n[SUCCESS] Successfully extracted {len(data)} records")
    print(f"[SUCCESS] Saved to: {output_path}")
    
    # Print summary
    if data:
        print(f"\nSample records:")
        for record in data[:5]:
            print(f"  {record['date']} | {record['store_name']} | Sales: RM {record['sales_amount']:,.2f} | Profit: RM {record['profit_amount']:,.2f}")
        if len(data) > 5:
            print(f"  ... and {len(data) - 5} more")

if __name__ == "__main__":
    main()

