"""
Investigate Excel File Structures
Compare Xilnex portal export vs Our API test export

Author: YONG WERN JIE
Date: October 28, 2025
"""

import pandas as pd
import openpyxl
from pathlib import Path

print("="*80)
print("EXCEL STRUCTURE INVESTIGATION")
print("="*80)
print()

# File paths
xilnex_file = r"C:\Users\MIS INTERN\Downloads\Daily Sales_28-10-2025 (Web).xlsx"
our_file = r"C:\Users\MIS INTERN\Downloads\Daily_Sales_API_Test_20181001_20181031.xlsx"

# Check files exist
if not Path(xilnex_file).exists():
    print(f"[ERROR] Xilnex file not found: {xilnex_file}")
    exit(1)

if not Path(our_file).exists():
    print(f"[ERROR] Our file not found: {our_file}")
    exit(1)

print("[1] XILNEX PORTAL EXPORT")
print("-"*80)

# Load Xilnex workbook
xilnex_wb = openpyxl.load_workbook(xilnex_file)
print(f"Sheets: {xilnex_wb.sheetnames}")
print()

# Analyze first sheet
xilnex_ws = xilnex_wb.active
print(f"Active Sheet: {xilnex_ws.title}")
print(f"Dimensions: {xilnex_ws.dimensions}")
print()

# Print first 30 rows to see structure
print("First 30 rows (raw):")
print("-"*80)
for i, row in enumerate(xilnex_ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    print(f"Row {i:2d}: {row}")
print()

# Try to read with pandas
print("Pandas read (default):")
try:
    xilnex_df = pd.read_excel(xilnex_file)
    print(f"Shape: {xilnex_df.shape}")
    print(f"Columns: {list(xilnex_df.columns)}")
    print()
    print("First 10 rows:")
    print(xilnex_df.head(10))
    print()
    print("Data types:")
    print(xilnex_df.dtypes)
except Exception as e:
    print(f"Error reading with pandas: {e}")

print()
print()

print("[2] OUR API TEST EXPORT")
print("-"*80)

# Load our workbook
our_wb = openpyxl.load_workbook(our_file)
print(f"Sheets: {our_wb.sheetnames}")
print()

# Analyze first sheet
our_ws = our_wb.active
print(f"Active Sheet: {our_ws.title}")
print(f"Dimensions: {our_ws.dimensions}")
print()

# Print first 30 rows
print("First 30 rows (raw):")
print("-"*80)
for i, row in enumerate(our_ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    print(f"Row {i:2d}: {row}")
print()

# Try to read with pandas
print("Pandas read (default):")
try:
    our_df = pd.read_excel(our_file)
    print(f"Shape: {our_df.shape}")
    print(f"Columns: {list(our_df.columns)}")
    print()
    print("First 10 rows:")
    print(our_df.head(10))
    print()
    print("Data types:")
    print(our_df.dtypes)
    print()
    print("Last 5 rows (to check for TOTAL row):")
    print(our_df.tail(5))
except Exception as e:
    print(f"Error reading with pandas: {e}")

print()
print()

print("[3] COMPARISON STRATEGY")
print("-"*80)
print()
print("Based on the structure analysis above, we need to:")
print("1. Identify header rows in Xilnex export")
print("2. Parse grouped/merged cells in Xilnex export")
print("3. Extract date totals from Xilnex (rows with 'Total')")
print("4. Aggregate our flat data by date")
print("5. Compare date totals")
print()
print("="*80)

