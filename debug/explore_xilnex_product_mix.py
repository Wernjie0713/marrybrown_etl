"""
Deep exploration of Xilnex Product Mix Report Excel structure.
This script analyzes the file structure to understand:
- Metadata headers
- Actual data location
- Column structure
- Merged cells
- Subtotal/total rows
- Hierarchical grouping patterns
"""

import pandas as pd
import openpyxl
from pathlib import Path

def explore_xilnex_excel(file_path):
    """
    Comprehensively analyze Xilnex Excel structure.
    """
    print("=" * 100)
    print("XILNEX PRODUCT MIX REPORT - STRUCTURE ANALYSIS")
    print("=" * 100)
    print(f"\nFile: {file_path}\n")
    
    # ============================================================================
    # STEP 1: Raw Excel Inspection using openpyxl (to see merged cells)
    # ============================================================================
    print("\n" + "=" * 100)
    print("STEP 1: RAW EXCEL STRUCTURE (using openpyxl)")
    print("=" * 100)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"\nWorksheet name: {ws.title}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")
    
    # Show merged cell ranges
    print(f"\n--- MERGED CELLS ---")
    if ws.merged_cells.ranges:
        print(f"Total merged cell ranges: {len(ws.merged_cells.ranges)}")
        for merged_range in list(ws.merged_cells.ranges)[:20]:  # Show first 20
            print(f"  {merged_range}")
    else:
        print("No merged cells found")
    
    # Show first 30 rows as-is
    print(f"\n--- FIRST 30 ROWS (RAW) ---")
    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(15, ws.max_column + 1)):  # Show first 15 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            # Show cell value and if it's part of a merged cell
            is_merged = any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges)
            row_values.append(f"{value} {'[M]' if is_merged else ''}")
        print(f"Row {row_idx:3d}: {' | '.join(row_values)}")
    
    # ============================================================================
    # STEP 2: Pandas Raw Load (with header=None)
    # ============================================================================
    print("\n" + "=" * 100)
    print("STEP 2: PANDAS RAW LOAD (header=None)")
    print("=" * 100)
    
    df_raw = pd.read_excel(file_path, header=None)
    print(f"\nShape: {df_raw.shape} (rows, columns)")
    print(f"\nFirst 30 rows:")
    print(df_raw.head(30).to_string())
    
    # ============================================================================
    # STEP 3: Identify Header Row
    # ============================================================================
    print("\n" + "=" * 100)
    print("STEP 3: IDENTIFY HEADER ROW")
    print("=" * 100)
    
    header_keywords = ['store', 'sales type', 'category', 'item name', 'item', 
                      'qty', 'quantity', 'cost', 'price', 'net sold', 'total']
    
    header_row_idx = None
    for idx, row in df_raw.iterrows():
        row_str = ' '.join([str(val).lower() for val in row.values if pd.notna(val)])
        matching_keywords = [kw for kw in header_keywords if kw in row_str]
        non_null_count = row.notna().sum()
        
        if len(matching_keywords) >= 3 and non_null_count >= 5:
            print(f"\nPotential header row at index {idx}:")
            print(f"  Non-null values: {non_null_count}")
            print(f"  Matching keywords: {matching_keywords}")
            print(f"  Row values: {row.tolist()}")
            if header_row_idx is None:
                header_row_idx = idx
    
    if header_row_idx is not None:
        print(f"\n[OK] Detected header row at index: {header_row_idx}")
    else:
        print("\n[FAILED] Could not detect header row")
    
    # ============================================================================
    # STEP 4: Load with Detected Header
    # ============================================================================
    print("\n" + "=" * 100)
    print("STEP 4: LOAD WITH DETECTED HEADER")
    print("=" * 100)
    
    if header_row_idx is not None:
        df = pd.read_excel(file_path, header=header_row_idx)
        print(f"\nShape after loading with header={header_row_idx}: {df.shape}")
        print(f"\nColumns: {df.columns.tolist()}")
        print(f"\nFirst 20 rows:")
        print(df.head(20).to_string())
        
        # ============================================================================
        # STEP 5: Analyze Column Patterns
        # ============================================================================
        print("\n" + "=" * 100)
        print("STEP 5: COLUMN ANALYSIS")
        print("=" * 100)
        
        for col in df.columns:
            non_null = df[col].notna().sum()
            null_pct = ((len(df) - non_null) / len(df)) * 100
            unique_vals = df[col].nunique()
            print(f"\nColumn: '{col}'")
            print(f"  Non-null: {non_null}/{len(df)} ({100-null_pct:.1f}%)")
            print(f"  Unique values: {unique_vals}")
            print(f"  Sample values: {df[col].dropna().head(5).tolist()}")
        
        # ============================================================================
        # STEP 6: Identify Subtotal/Total Rows
        # ============================================================================
        print("\n" + "=" * 100)
        print("STEP 6: IDENTIFY SUBTOTAL/TOTAL ROWS")
        print("=" * 100)
        
        # Check Item Name column for "Total" keyword
        item_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if 'item' in col_lower or 'name' in col_lower or 'product' in col_lower:
                item_col = col
                break
        
        if item_col:
            print(f"\nAnalyzing column: '{item_col}'")
            total_rows = df[df[item_col].astype(str).str.contains('total', case=False, na=False)]
            print(f"\nFound {len(total_rows)} rows containing 'Total':")
            print(total_rows[[item_col] + [c for c in df.columns if c != item_col][:5]].to_string())
        else:
            print("\nCould not identify Item Name column")
        
        # ============================================================================
        # STEP 7: Analyze Hierarchical Structure (Store -> Sales Type -> Category)
        # ============================================================================
        print("\n" + "=" * 100)
        print("STEP 7: HIERARCHICAL STRUCTURE ANALYSIS")
        print("=" * 100)
        
        store_col = None
        sales_type_col = None
        category_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if 'store' in col_lower:
                store_col = col
            elif 'sales' in col_lower and 'type' in col_lower:
                sales_type_col = col
            elif 'category' in col_lower:
                category_col = col
        
        print(f"\nIdentified hierarchical columns:")
        print(f"  Store: {store_col}")
        print(f"  Sales Type: {sales_type_col}")
        print(f"  Category: {category_col}")
        
        if store_col:
            print(f"\n--- Store Column Analysis ---")
            print(f"Unique stores: {df[store_col].nunique()}")
            print(f"Unique values: {df[store_col].unique()[:10]}")  # Show first 10
            print(f"Value counts:")
            print(df[store_col].value_counts().head(10))
        
        if sales_type_col:
            print(f"\n--- Sales Type Column Analysis ---")
            print(f"Unique sales types: {df[sales_type_col].nunique()}")
            print(f"Unique values: {df[sales_type_col].unique()[:10]}")
            print(f"Value counts:")
            print(df[sales_type_col].value_counts().head(10))
        
        if category_col:
            print(f"\n--- Category Column Analysis ---")
            print(f"Unique categories: {df[category_col].nunique()}")
            print(f"Unique values: {df[category_col].unique()[:20]}")  # Show first 20
            print(f"Value counts:")
            print(df[category_col].value_counts().head(20))
        
        # ============================================================================
        # STEP 8: Sample Data Rows (Non-Total Rows)
        # ============================================================================
        print("\n" + "=" * 100)
        print("STEP 8: SAMPLE ACTUAL DATA ROWS (excluding totals)")
        print("=" * 100)
        
        if item_col:
            data_rows = df[~df[item_col].astype(str).str.contains('total', case=False, na=False)]
            data_rows = data_rows.dropna(how='all')
            print(f"\nTotal data rows (non-totals): {len(data_rows)}")
            print(f"\nFirst 10 data rows:")
            print(data_rows.head(10).to_string())
        
    # ============================================================================
    # SUMMARY & RECOMMENDATIONS
    # ============================================================================
    print("\n" + "=" * 100)
    print("SUMMARY & RECOMMENDATIONS")
    print("=" * 100)
    
    print("\n1. METADATA ROWS:")
    print(f"   - Rows before header: {header_row_idx if header_row_idx else 'Unknown'}")
    
    print("\n2. HEADER ROW:")
    print(f"   - Located at index: {header_row_idx if header_row_idx else 'Not found'}")
    
    print("\n3. MERGED CELLS:")
    print(f"   - Total merged ranges: {len(ws.merged_cells.ranges) if ws.merged_cells.ranges else 0}")
    print("   - This indicates hierarchical grouping (Store/Sales Type/Category)")
    
    print("\n4. DATA STRUCTURE:")
    if header_row_idx is not None and item_col:
        print(f"   - Total rows loaded: {len(df)}")
        print(f"   - Actual data rows: {len(data_rows) if 'data_rows' in locals() else 'Unknown'}")
        print(f"   - Subtotal/Total rows: {len(total_rows) if 'total_rows' in locals() else 'Unknown'}")
    
    print("\n5. RECOMMENDED PARSING STRATEGY:")
    print("   a) Skip first N rows (metadata)")
    print("   b) Use detected header row")
    print("   c) Filter out rows where Item Name contains 'Total'")
    print("   d) Handle merged cells by forward-filling Store/Sales Type/Category")
    print("   e) Clean empty rows (dropna(how='all'))")

if __name__ == "__main__":
    xilnex_file = r"C:\Users\MIS INTERN\Downloads\Product Mix Report_27-10-2025 (Web).xlsx"
    
    if not Path(xilnex_file).exists():
        print(f"ERROR: File not found: {xilnex_file}")
        print("\nPlease update the file path in the script.")
    else:
        explore_xilnex_excel(xilnex_file)
        print("\n" + "=" * 100)
        print("ANALYSIS COMPLETE")
        print("=" * 100)

